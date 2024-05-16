from typing import List
from openpyxl import Workbook
from sqlalchemy.orm import Session
from src.models.db.first_stage_analysis import FirstStageAnalysisModel
from src.models.db.currency_base_info import CurrencyBaseInfoModel
from src.models.db.setor import Setor
from src.models.db.rel_setor_currency_base_info import SetorCurrencyBaseInfo

class WorkbookService:
    def __init__(self, session: Session):
        self.session = session

    def create_workbook(self, headers: List[str]) -> Workbook:
        workbook = Workbook()
        worksheet = workbook.active
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)
        return workbook

    def fill_workbook(self, workbook: Workbook, headers: List[str]) -> Workbook:
        worksheet = workbook.active
        data = self.session.query(FirstStageAnalysisModel).join(
            CurrencyBaseInfoModel,
            FirstStageAnalysisModel.uuid_currency == CurrencyBaseInfoModel.uuid
        )

        header_to_model_attr = {
            #"CATEGORY": ,
            "SYMBOL": lambda item: item.currency.symbol if item.currency else "N/A",
            "RANKING": "ranking",
            "MARKET CAP": "market_cap",
            "INCREASE DATE": "increase_date",
            "% WEEK INCREASE": "week_increase_percentage",
            "WEEK CLOSING PRICE": "closing_price",
            "WEEK OPEN PRICE": "open_price",
            "WEEK CLOSING PRICE > EMA8(w)": "ema8_less_close",
            "EMA8 > WEEK OPEN PRICE": "ema8_greater_open",
            "EMAs ALIGNED": "ema_aligned",
            "INCREASE VOLUME(d) DATE": "increase_volume_day",
            "INCREASE VOLUME(d)": "week_increase_volume",
            "DAY BEFORE VOLUME(d)": "day_before_volume",
            "% VOLUME/VOLUME DAY BEFORE": "volumes_relation",
            "VOLUME > 200%": "expressive_volume_increase",
            "BUY SIGNAL": "buying_signal",
            #"1 YEAR": "year_variation_per",
            #"180 DAYS": "semester_variation_per",
            #"90 DAYS": "quarter_variation_per",
            #"30 DAYS": "month_variation_per",
            #"7 DAYS": "week_variation_per"
        }
        
        for row_idx, item in enumerate(data, start=2):
            for header in headers:
                col_idx = headers.index(header) + 1
                model_attr = header_to_model_attr.get(header)
                
                value = model_attr(item) if callable(model_attr) else getattr(item, model_attr, "N/A") if model_attr is not None else "N/A"
                worksheet.cell(row=row_idx, column=col_idx, value=value)
        return workbook
